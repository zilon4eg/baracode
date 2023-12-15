from barcode import Code128
from barcode.writer import SVGWriter
from openpyxl import load_workbook
from PyPDF2 import PdfWriter
from reportlab.graphics import renderPDF, renderPM
from svglib.svglib import svg2rlg
import os


def generate_svg(code, path):
    options = {
                'module_width': 0.3,
                'module_height': 10.0,
                'font_size': 10
            }
    writer = SVGWriter()
    writer.set_options(options)
    barcode = Code128(str(code), writer)
    barcode.save(f'{path}\\{code}', options)


def svg_to_pdf(code, path):
    drawing = svg2rlg(f'{path}\\{code}.svg')
    renderPDF.drawToFile(drawing, f'{path}\\{code}.pdf')


def create_pdf():
    merger = PdfWriter()
    count = 0
    for cell in col_a:
        count += 1
        generate_svg(cell.value, mainpath)
        svg_to_pdf(cell.value, mainpath)
        os.remove(f'{mainpath}\\{cell.value}.svg')
        merger.append(f'{mainpath}\\{cell.value}.pdf')
        os.remove(f'{mainpath}\\{cell.value}.pdf')
        print(f'{ws.title} => {count} of {len(col_a)} complete')
    merger.write(f'{mainpath}\\{ws.title}.pdf')
    merger.close()


if __name__ == '__main__':
    mainpath = r'C:\Users\hdd50\Desktop\baracode2'
    wb = load_workbook(f'{mainpath}\\Accred_09-10_допка.xlsx')

    for ws in wb:
        col_a = ws['A']
        create_pdf()
